#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 차트 자동 삽입 시스템

키워드 트렌드 분석 결과를 Excel 시트에 차트로 자동 삽입합니다.
"""

import sys
import os

# UTF-8 출력 설정 (Windows)
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass

import logging
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from pathlib import Path
from io import BytesIO

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # GUI 없이 이미지 생성

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import (
    BarChart, LineChart, PieChart, Reference,
    Series, ScatterChart, BubbleChart
)
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================================
# 설정 및 로깅
# ============================================================================

class ExcelChartConfig:
    """Excel 차트 설정"""

    # 차트 설정
    CHART_WIDTH: int = 15
    CHART_HEIGHT: int = 10
    CHART_DPI: int = 150

    # 색상 테마
    COLORS: List[str] = [
        '#6B9BD1', '#ED7D31', '#A5A5A5', '#FFC000',
        '#4472C4', '#70AD47', '#FF0000', '#002060'
    ]

    # 폰트 설정
    FONT_NAME: str = '맑은 고딕'
    TITLE_FONT_SIZE: int = 14
    LABEL_FONT_SIZE: int = 10

    # 여백 설정
    CHART_START_ROW: int = 2
    CHART_START_COL: int = 2


def setup_logging(name: str = "excel_chart") -> logging.Logger:
    """로그 설정"""
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)

    if not logger.handlers:
        handler = logging.StreamHandler()
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        handler.setFormatter(formatter)
        logger.addHandler(handler)

    return logger


# ============================================================================
# Excel 차트 삽입 클래스
# ============================================================================

class ExcelChartInserter:
    """Excel 차트 자동 삽입기"""

    def __init__(self, config: ExcelChartConfig = None):
        """
        차트 삽입기 초기화

        Args:
            config: 차트 설정
        """
        self.config = config or ExcelChartConfig()
        self.logger = setup_logging()

        # matplotlib 한글 폰트 설정
        self._setup_matplotlib_font()

    def _setup_matplotlib_font(self):
        """matplotlib 한글 폰트 설정"""
        try:
            plt.rcParams['font.family'] = self.config.FONT_NAME
            plt.rcParams['axes.unicode_minus'] = False
        except:
            self.logger.warning("한글 폰트 설정 실패")

    def insert_keyword_frequency_chart(self, excel_path: str,
                                      keyword_data: List[Tuple[str, int]],
                                      sheet_name: str = "키워드 분석",
                                      position: str = "H2") -> bool:
        """
        키워드 빈도 차트 삽입

        Args:
            excel_path: Excel 파일 경로
            keyword_data: [(키워드, 빈도수), ...] 형태의 데이터
            sheet_name: 시트 이름
            position: 차트 위치 (예: "H2")

        Returns:
            삽입 성공 여부
        """
        try:
            # 워크북 열기
            wb = load_workbook(excel_path)

            # 시트 선택 또는 생성
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            # matplotlib으로 차트 생성
            fig, ax = plt.subplots(figsize=(self.config.CHART_WIDTH, self.config.CHART_HEIGHT))

            keywords = [kw[0] for kw in keyword_data[:20]]
            frequencies = [kw[1] for kw in keyword_data[:20]]

            # 막대 그래프
            bars = ax.barh(keywords, frequencies, color=self.config.COLORS[0])
            ax.set_xlabel('빈도수', fontsize=self.config.LABEL_FONT_SIZE)
            ax.set_ylabel('키워드', fontsize=self.config.LABEL_FONT_SIZE)
            ax.set_title('키워드 등장 빈도', fontsize=self.config.TITLE_FONT_SIZE, fontweight='bold')
            ax.invert_yaxis()

            # 값 표시
            for bar, freq in zip(bars, frequencies):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2,
                       f'{int(freq)}', ha='left', va='center',
                       fontsize=self.config.LABEL_FONT_SIZE - 2)

            plt.tight_layout()

            # 이미지로 저장
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=self.config.CHART_DPI,
                       bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()

            # Excel에 삽입
            img = Image(img_buffer)
            img.width = 800
            img.height = 600

            # 위치 계산
            col = int(position[1:-1]) - 1 if position[-1].isdigit() else 0
            row = int(''.join(filter(str.isdigit, position)))

            ws.add_image(img, position)

            # 저장
            wb.save(excel_path)
            wb.close()

            self.logger.info(f"키워드 빈도 차트 삽입 완료: {sheet_name}!{position}")

            return True

        except Exception as e:
            self.logger.error(f"키워드 빈도 차트 삽입 실패: {e}")
            return False

    def insert_trend_line_chart(self, excel_path: str,
                               trend_df: pd.DataFrame,
                               sheet_name: str = "트렌드 분석",
                               position: str = "H2") -> bool:
        """
        트렌드 라인 차트 삽입

        Args:
            excel_path: Excel 파일 경로
            trend_df: 트렌드 데이터프레임
            sheet_name: 시트 이름
            position: 차트 위치

        Returns:
            삽입 성공 여부
        """
        try:
            wb = load_workbook(excel_path)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            # 상위 10개 키워드 선택
            top_keywords = trend_df.groupby('keyword')['frequency'].sum().nlargest(10).index

            # matplotlib 차트 생성
            fig, ax = plt.subplots(figsize=(self.config.CHART_WIDTH, self.config.CHART_HEIGHT))

            for i, keyword in enumerate(top_keywords):
                keyword_data = trend_df[trend_df['keyword'] == keyword].sort_values('date')
                dates = pd.to_datetime(keyword_data['date'])
                frequencies = keyword_data['frequency'].values

                ax.plot(dates, frequencies, marker='o', label=keyword,
                       color=self.config.COLORS[i % len(self.config.COLORS)][-6:],
                       linewidth=2, markersize=6)

            ax.set_xlabel('날짜', fontsize=self.config.LABEL_FONT_SIZE)
            ax.set_ylabel('빈도수', fontsize=self.config.LABEL_FONT_SIZE)
            ax.set_title('키워드 트렌드 추이', fontsize=self.config.TITLE_FONT_SIZE, fontweight='bold')
            ax.legend(loc='best', fontsize=self.config.LABEL_FONT_SIZE - 2)
            ax.grid(True, alpha=0.3)

            plt.xticks(rotation=45)
            plt.tight_layout()

            # 이미지 저장
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=self.config.CHART_DPI,
                       bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()

            # Excel 삽입
            img = Image(img_buffer)
            img.width = 800
            img.height = 600
            ws.add_image(img, position)

            wb.save(excel_path)
            wb.close()

            self.logger.info(f"트렌드 라인 차트 삽입 완료: {sheet_name}!{position}")

            return True

        except Exception as e:
            self.logger.error(f"트렌드 라인 차트 삽입 실패: {e}")
            return False

    def insert_growth_heatmap(self, excel_path: str,
                            growth_df: pd.DataFrame,
                            sheet_name: str = "성장률 분석",
                            position: str = "H2") -> bool:
        """
        성장률 히트맵 차트 삽입

        Args:
            excel_path: Excel 파일 경로
            growth_df: 성장률 데이터프레임
            sheet_name: 시트 이름
            position: 차트 위치

        Returns:
            삽입 성공 여부
        """
        try:
            wb = load_workbook(excel_path)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            # 피벗 테이블 생성
            pivot_df = growth_df.pivot(index='keyword', columns='date', values='growth_rate')

            # matplotlib 히트맵
            fig, ax = plt.subplots(figsize=(self.config.CHART_WIDTH, self.config.CHART_HEIGHT))

            im = ax.imshow(pivot_df.values, cmap='RdYlGn', aspect='auto')

            # 축 설정
            ax.set_xticks(range(len(pivot_df.columns)))
            ax.set_yticks(range(len(pivot_df.index)))
            ax.set_xticklabels([str(d).split()[0] for d in pivot_df.columns], rotation=45)
            ax.set_yticklabels(pivot_df.index)

            # 값 표시
            for i in range(len(pivot_df.index)):
                for j in range(len(pivot_df.columns)):
                    value = pivot_df.values[i, j]
                    text_color = 'white' if abs(value) > 50 else 'black'
                    ax.text(j, i, f'{value:.0f}%', ha='center', va='center',
                           color=text_color, fontsize=self.config.LABEL_FONT_SIZE - 2)

            ax.set_xlabel('날짜', fontsize=self.config.LABEL_FONT_SIZE)
            ax.set_ylabel('키워드', fontsize=self.config.LABEL_FONT_SIZE)
            ax.set_title('키워드 성장률 히트맵', fontsize=self.config.TITLE_FONT_SIZE, fontweight='bold')

            # 컬러바
            cbar = plt.colorbar(im, ax=ax)
            cbar.set_label('성장률 (%)', rotation=270, labelpad=20)

            plt.tight_layout()

            # 이미지 저장
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=self.config.CHART_DPI,
                       bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()

            # Excel 삽입
            img = Image(img_buffer)
            img.width = 800
            img.height = 600
            ws.add_image(img, position)

            wb.save(excel_path)
            wb.close()

            self.logger.info(f"성장률 히트맵 삽입 완료: {sheet_name}!{position}")

            return True

        except Exception as e:
            self.logger.error(f"성장률 히트맵 삽입 실패: {e}")
            return False

    def insert_native_excel_chart(self, excel_path: str,
                                 data_df: pd.DataFrame,
                                 chart_type: str = "bar",
                                 sheet_name: str = "Excel 차트",
                                 position: str = "A2") -> bool:
        """
        Excel 네이티브 차트 삽입 (openpyxl)

        Args:
            excel_path: Excel 파일 경로
            data_df: 데이터프레임
            chart_type: 차트 타입 ("bar", "line", "pie")
            sheet_name: 시트 이름
            position: 데이터 시작 위치

        Returns:
            삽입 성공 여부
        """
        try:
            wb = load_workbook(excel_path)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            # 데이터 시트에 쓰기
            ws.append(['키워드'] + list(data_df.columns[1:]))

            for _, row in data_df.iterrows():
                ws.append(list(row))

            # 차트 생성
            if chart_type == "bar":
                chart = self._create_bar_chart(ws, data_df, position)
            elif chart_type == "line":
                chart = self._create_line_chart(ws, data_df, position)
            elif chart_type == "pie":
                chart = self._create_pie_chart(ws, data_df, position)
            else:
                raise ValueError(f"지원하지 않는 차트 타입: {chart_type}")

            # 차트 추가
            chart_pos = self._get_chart_position(position, rows_offset=len(data_df) + 5)
            ws.add_chart(chart, chart_pos)

            wb.save(excel_path)
            wb.close()

            self.logger.info(f"Excel 네이티브 {chart_type} 차트 삽입 완료")

            return True

        except Exception as e:
            self.logger.error(f"Excel 네이티브 차트 삽입 실패: {e}")
            return False

    def _create_bar_chart(self, ws, data_df: pd.DataFrame, position: str) -> BarChart:
        """막대 차트 생성"""
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "키워드 빈도 분석"
        chart.y_axis.title = '빈도수'
        chart.x_axis.title = '키워드'

        # 데이터 참조
        data_start_col = 2  # B열
        data_end_col = data_start_col + len(data_df.columns) - 1
        data_start_row = int(position[1:]) if position[1:].isdigit() else 2
        data_end_row = data_start_row + len(data_df)

        data_ref = Reference(ws, min_col=data_start_col, min_row=data_start_row,
                            max_col=data_end_col, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row+1,
                            max_row=data_end_row)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        return chart

    def _create_line_chart(self, ws, data_df: pd.DataFrame, position: str) -> LineChart:
        """라인 차트 생성"""
        chart = LineChart()
        chart.style = 10
        chart.title = "키워드 트렌드"
        chart.y_axis.title = '빈도수'
        chart.x_axis.title = '기간'

        data_start_col = 2
        data_end_col = data_start_col + len(data_df.columns) - 1
        data_start_row = int(position[1:]) if position[1:].isdigit() else 2
        data_end_row = data_start_row + len(data_df)

        data_ref = Reference(ws, min_col=data_start_col, min_row=data_start_row,
                            max_col=data_end_col, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row+1,
                            max_row=data_end_row)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        return chart

    def _create_pie_chart(self, ws, data_df: pd.DataFrame, position: str) -> PieChart:
        """파이 차트 생성"""
        chart = PieChart()
        chart.title = "키워드 비율"

        data_start_col = 2
        data_start_row = int(position[1:]) if position[1:].isdigit() else 2
        data_end_row = data_start_row + len(data_df)

        data_ref = Reference(ws, min_col=data_start_col, min_row=data_start_row,
                            max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row+1,
                            max_row=data_end_row)

        chart.add_data(data_ref)
        chart.set_categories(cats_ref)

        return chart

    def _get_chart_position(self, data_position: str, rows_offset: int = 5) -> str:
        """차트 위치 계산"""
        col = data_position[0]
        row = int(data_position[1:]) + rows_offset
        return f"{col}{row}"

    def insert_network_diagram(self, excel_path: str,
                              network_graph,
                              sheet_name: str = "네트워크 분석",
                              position: str = "A2") -> bool:
        """
        네트워크 다이어그램 삽입

        Args:
            excel_path: Excel 파일 경로
            network_graph: NetworkX 그래프 객체
            sheet_name: 시트 이름
            position: 차트 위치

        Returns:
            삽입 성공 여부
        """
        try:
            import networkx as nx

            wb = load_workbook(excel_path)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            # matplotlib 네트워크 플롯
            fig, ax = plt.subplots(figsize=(self.config.CHART_WIDTH, self.config.CHART_HEIGHT))

            # 레이아웃
            pos = nx.spring_layout(network_graph, k=1, iterations=50)

            # 노드 크기 (빈도수 기반)
            node_sizes = [network_graph.nodes[node].get('frequency', 1) * 100
                         for node in network_graph.nodes()]

            # 노드 색상 (연결 중심성 기반)
            centrality = nx.degree_centrality(network_graph)
            node_colors = [centrality[node] for node in network_graph.nodes()]

            # 그리기
            nx.draw_networkx_nodes(network_graph, pos, node_size=node_sizes,
                                  node_color=node_colors, cmap='YlOrRd', ax=ax)
            nx.draw_networkx_edges(network_graph, pos, alpha=0.3, width=0.5, ax=ax)
            nx.draw_networkx_labels(network_graph, pos, font_size=8,
                                   font_family=self.config.FONT_NAME, ax=ax)

            ax.set_title('키워드 연관 네트워크', fontsize=self.config.TITLE_FONT_SIZE,
                        fontweight='bold')
            ax.axis('off')

            plt.tight_layout()

            # 이미지 저장
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=self.config.CHART_DPI,
                       bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()

            # Excel 삽입
            img = Image(img_buffer)
            img.width = 1000
            img.height = 800
            ws.add_image(img, position)

            wb.save(excel_path)
            wb.close()

            self.logger.info(f"네트워크 다이어그램 삽입 완료: {sheet_name}!{position}")

            return True

        except Exception as e:
            self.logger.error(f"네트워크 다이어그램 삽입 실패: {e}")
            return False

    def insert_wordcloud(self, excel_path: str,
                        word_freq: Dict[str, int],
                        sheet_name: str = "워드클라우드",
                        position: str = "A2") -> bool:
        """
        워드클라우드 삽입

        Args:
            excel_path: Excel 파일 경로
            word_freq: {단어: 빈도수} 형태의 딕셔너리
            sheet_name: 시트 이름
            position: 차트 위치

        Returns:
            삽입 성공 여부
        """
        try:
            from wordcloud import WordCloud

            wb = load_workbook(excel_path)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            # 워드클라우드 생성
            wc = WordCloud(width=1200, height=800,
                          background_color='white',
                          colormap='viridis',
                          font_path='C:/Windows/Fonts/malgun.ttf',
                          max_words=100)

            wc.generate_from_frequencies(word_freq)

            # matplotlib로 시각화
            fig, ax = plt.subplots(figsize=(self.config.CHART_WIDTH, self.config.CHART_HEIGHT))
            ax.imshow(wc, interpolation='bilinear')
            ax.axis('off')
            ax.set_title('키워드 워드클라우드', fontsize=self.config.TITLE_FONT_SIZE,
                       fontweight='bold')

            plt.tight_layout()

            # 이미지 저장
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=self.config.CHART_DPI,
                       bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()

            # Excel 삽입
            img = Image(img_buffer)
            img.width = 1000
            img.height = 800
            ws.add_image(img, position)

            wb.save(excel_path)
            wb.close()

            self.logger.info(f"워드클라우드 삽입 완료: {sheet_name}!{position}")

            return True

        except Exception as e:
            self.logger.error(f"워드클라우드 삽입 실패: {e}")
            return False


# ============================================================================
# 통합 Excel 리포트 생성기
# ============================================================================

class ExcelReportGenerator:
    """Excel 분석 리포트 생성기"""

    def __init__(self, config: ExcelChartConfig = None):
        """
        리포트 생성기 초기화

        Args:
            config: 차트 설정
        """
        self.config = config or ExcelChartConfig()
        self.inserter = ExcelChartInserter(config)
        self.logger = setup_logging()

    def generate_comprehensive_report(self,
                                     excel_path: str,
                                     trend_df: pd.DataFrame,
                                     growth_df: pd.DataFrame,
                                     keyword_freq: List[Tuple[str, int]],
                                     network_graph=None,
                                     word_freq: Dict[str, int] = None) -> Dict[str, bool]:
        """
        종합 Excel 리포트 생성

        Args:
            excel_path: Excel 파일 경로
            trend_df: 트렌드 데이터프레임
            growth_df: 성장률 데이터프레임
            keyword_freq: 키워드 빈도 리스트
            network_graph: 네트워크 그래프 (선택사항)
            word_freq: 워드클라우드 데이터 (선택사항)

        Returns:
            각 차트별 삽입 결과
        """
        results = {}

        # 1. 키워드 빈도 차트
        results['frequency_chart'] = self.inserter.insert_keyword_frequency_chart(
            excel_path, keyword_freq, sheet_name="키워드 빈도", position="H2"
        )

        # 2. 트렌드 라인 차트
        results['trend_chart'] = self.inserter.insert_trend_line_chart(
            excel_path, trend_df, sheet_name="트렌드 분석", position="H2"
        )

        # 3. 성장률 히트맵
        if not growth_df.empty:
            results['heatmap'] = self.inserter.insert_growth_heatmap(
                excel_path, growth_df, sheet_name="성장률 분석", position="H2"
            )

        # 4. 네트워크 다이어그램
        if network_graph is not None:
            results['network'] = self.inserter.insert_network_diagram(
                excel_path, network_graph, sheet_name="네트워크 분석", position="A2"
            )

        # 5. 워드클라우드
        if word_freq is not None:
            results['wordcloud'] = self.inserter.insert_wordcloud(
                excel_path, word_freq, sheet_name="워드클라우드", position="A2"
            )

        success_count = sum(1 for v in results.values() if v)
        self.logger.info(f"Excel 리포트 생성 완료: {success_count}/{len(results)}개 차트")

        return results


# ============================================================================
# 메인 실행
# ============================================================================

def main():
    """메인 실행 함수"""
    print("Excel 차트 자동 삽입 시스템 테스트")

    # 테스트 데이터 생성
    test_data = {
        'keyword': ['AI', '블록체인', '머신러닝', '딥러닝', '데이터'],
        'frequency': [100, 80, 60, 40, 20]
    }

    # 리포트 생성기 초기화
    generator = ExcelReportGenerator()

    print("Excel 차트 삽입 기능이 준비되었습니다.")
    print("사용법: ExcelReportGenerator 클래스의 메서드를 호출하세요.")


if __name__ == "__main__":
    main()
